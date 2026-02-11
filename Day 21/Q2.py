import pandas as pd
from openpyxl import load_workbook, Workbook

df = pd.read_excel("sales_data.xlsx", sheet_name="2025")
df["Total"] = df["Quantity"] * df["Price"]
df.to_excel("sales_summary.xlsx", index=False)

wb = load_workbook("sales_data.xlsx")
ws = wb["2025"]

new_wb = Workbook()
new_ws = new_wb.active

headers = [cell.value for cell in ws[1]]
headers.append("Total")
new_ws.append(headers)

for row in ws.iter_rows(min_row=2, values_only=True):
    product, quantity, price = row
    total = quantity * price
    new_ws.append([product, quantity, price, total])

new_wb.save("sales_summary_openpyxl.xlsx")
